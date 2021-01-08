# Alt Z Toggle word wrap
import requests
from bs4 import BeautifulSoup
import re # To find part of string
import openpyxl
import pykakasi # to translate hiragana -> furigana
import cfscrape
import os
from os import listdir

kks = pykakasi.kakasi()
excel_name = 'new.xlsx'

http_add_begin = "https://www.japandict.com/voice/read?text="
http_add_content = "&outputFormat=ogg_vorbis&jwt="
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
# header = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,'referer':'https://www.google.com/'}

# headers = {
#     "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:81.0) Gecko/20100101 Firefox/81.0" 
# }

def download_sound(http_to_sound,row,word_furigana_name):
    with open("/home/hung/Documents/PythonVisual/python-anki/sound/"+str(row) + "_" + word_furigana_name, "wb") as f:
        r = requests.get(http_to_sound, headers=headers)
        f.write(r.content)
        print("write: "+word_furigana_name+" done")

def write_new_excel_file(excel_name,row,word,sound_list):
    wb = openpyxl.load_workbook(excel_name)
    ws = wb['sound_list']
    ws.cell(row,1).value = word
    column = 2
    for sound in sound_list:
        ws.cell(row,column).value = sound
        column += 1
    wb.save(excel_name)

def get_word_from_dict(word,row):
    print(str(row) + ": get :"+word)
    r = requests.get("https://www.japandict.com/"+word+"?lang=eng",headers=headers)
    soup = BeautifulSoup(r.content,'lxml')
    class_sound = soup.find_all("a",class_="btn btn-secondary-outline p-b-0 p-t-0 play-reading-btn")
    sound_list = {''}
    if class_sound  is not None:
        for class_sound_subset in class_sound:
            sound_http = re.split('"', class_sound_subset['data-reading'])
            if sound_http is not None:
                http_to_sound = http_add_begin +  sound_http[3] + http_add_content + sound_http[5]
                word_furigana = kks.convert(sound_http[3])
                for word_furigana_subset in word_furigana:
                    word_furigana_name = word_furigana_subset['hepburn']
                download_sound(http_to_sound,row,word_furigana_name)
                sound_list.add(word_furigana_name)
                print(word+" has")
            else:
                print(word+" not has")
                return False
    else:
        print(word+" not has")
        return False
    # write_new_excel_file(excel_name,row,word,sound_list)

# get_word_from_dict("あきれる",1)
def get_word_kantan(word, name):
    path = f"https://kantan.vn//?keyword={word}"
    r = requests.get(path,headers=headers)
    soup = BeautifulSoup(r.content,'lxml')
    print(soup)
    spans = soup.find_all('span', {'class' : 'nvmn-meaning'})
    # mydivs = soup.findAll("span", {"class": "nvmn-meaning"})
    # class_sound = soup.find_all("span",class_="nvmn-meaning")
    print(spans)

# get_word_kantan("辞書","hung")


def read_excel():
    wb = openpyxl.load_workbook('a.xlsx')
    ws = wb["n2"]
    for row in range(1,ws.max_row):
        word = str(ws.cell(row,1).value)
        print(word)
        # if get_word_from_dict(word,row) == True:
        #     ws.cell(row,5).value = row
        #     print(ws.cell(row,1).value) 
    wb.save('a.xlsx')
    print("done")

# read_excel()

def folder_file_to_Excel():
    path = "/home/hung/Documents/PythonVisual/python-anki/sound"
    wb = openpyxl.load_workbook("hung.xlsx")
    ws = wb['sound_list']
    i = 0
    for file in listdir(path):
        word = split_word_by_underscore(file)
        number_temp = int(word[0])
        word = str(word[1])
        i = i+1
        ws.cell(i,1).value = word
        ws.cell(i,2).value = number_temp
    wb.save(excel_name)

def rename_sound_in_excel():
    path = "/home/hung/Desktop/a.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb['a']
    for row in range(1,ws.max_row):
        word = str(ws.cell(row,3).value)
        temp = word.replace("[sound:","")
        temp = temp.replace(".mp3]","")
        word_furigana = kks.convert(temp)
        for word_furigana_subset in word_furigana:
            result = word_furigana_subset['hepburn']
            result = "[sound:"+result+".mp3]"
            ws.cell(row,3).value = result
            print(word)
            print(result)
    wb.save(path)

rename_sound_in_excel()


def rename_file():
    path = "/home/hung/Desktop/from"
    for file in listdir(path):
        file1 = file.replace(".mp3","")
        word_furigana = kks.convert(file1)
        for word_furigana_subset in word_furigana:
            result = word_furigana_subset['hepburn']
            result = path +"/"+ result+".mp3"
            file = path + "/" + file
            print(file)
            print(result)
            if(os.path.isfile(file)):
                os.rename(file, result)

# rename_file()


def remain_file():
    path = "/home/hung/Documents/PythonVisual/python-anki/sound_22_10"
    for file in listdir(path):
        # os.rename(path+"/"+file,path+"/"+file+"_sound.mp3")
        if 'sound' in file:
            os.rename(path+"/"+file,path+"/"+file+".mp3")
            # print(file+".mp3")
        # print(file)
        # word = split_word_by_underscore(file)
        # print(word)
        # word = str(word[1])
        # os.rename(path+"/"+file,path+"/"+word)

# remain_file()

def split_word_by_underscore(word):
    word_list = word.split("_")
    return word_list

# remain_file()

# folder_file_to_Excel()
